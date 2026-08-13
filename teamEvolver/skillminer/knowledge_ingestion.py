"""Normalize uploaded knowledge documents into UTF-8 Markdown.

The mining pipeline and its sample-package validator intentionally consume
plain text files.  Binary Office/PDF inputs are therefore converted at the
ingestion boundary instead of being handed to Hermes directly.
"""

from __future__ import annotations

import contextlib
import hashlib
import io
import json
import os
import re
import threading
import time as time_module
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Iterator

SUPPORTED_KNOWLEDGE_SUFFIXES = {
    ".md",
    ".markdown",
    ".txt",
    ".docx",
    ".xlsx",
    ".pdf",
}

MAX_ARCHIVE_FILES = 10_000
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_NORMALIZED_DOCUMENT_CHARS = 5_000_000
MAX_SPREADSHEET_CELLS = 200_000
MAX_PDF_PAGES = 2_000
INGESTION_STATE_FILENAME = ".ingestion-state.json"

_SOURCE_LOCKS: dict[str, threading.Lock] = {}
_SOURCE_LOCKS_GUARD = threading.Lock()


@dataclass(frozen=True)
class NormalizedKnowledge:
    markdown: str
    source_format: str
    source_encoding: str


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _state_path(project_root: Path | str, source_name: str) -> Path:
    project_root = Path(project_root).resolve()
    originals_root = (project_root / ".knowledge_originals").resolve()
    source_root = (originals_root / str(source_name)).resolve()
    if source_root.parent != originals_root or not source_root.name or source_root.name.startswith("."):
        raise ValueError("知识源状态路径无效")
    return source_root / INGESTION_STATE_FILENAME


def _default_state(source_name: str, has_documents: bool) -> dict[str, object]:
    status = "ready" if has_documents else "empty"
    return {
        "schema_version": 1,
        "source_path": f"data/{source_name}",
        "batch_id": "",
        "status": status,
        "stage": "complete" if has_documents else "idle",
        "progress": 100 if has_documents else 0,
        "processed_files": 0,
        "total_files": 0,
        "current_file": "",
        "error": "",
        "started_at": "",
        "updated_at": "",
        "finished_at": "",
        "pending_outputs": [],
    }


def read_ingestion_state(
    project_root: Path | str,
    source_name: str,
    *,
    has_documents: bool,
) -> dict[str, object]:
    """Read the persistent state, with safe defaults for legacy data sources."""
    path = _state_path(project_root, source_name)
    if not path.is_file():
        return _default_state(source_name, has_documents)
    try:
        state = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(state, dict) or state.get("status") not in {
            "empty", "processing", "ready", "failed",
        }:
            raise ValueError("invalid state")
    except (OSError, ValueError, json.JSONDecodeError):
        state = _default_state(source_name, has_documents)
        state.update({
            "status": "failed",
            "stage": "failed",
            "progress": 0,
            "error": "知识源处理状态损坏，请重新上传文件",
        })
        return state

    result = _default_state(source_name, has_documents)
    result.update(state)
    result["source_path"] = f"data/{source_name}"
    result["progress"] = max(0, min(100, int(result.get("progress") or 0)))
    result["processed_files"] = max(0, int(result.get("processed_files") or 0))
    result["total_files"] = max(0, int(result.get("total_files") or 0))
    if result["status"] == "ready" and not has_documents:
        result.update({"status": "empty", "stage": "idle", "progress": 0})
    return result


def write_ingestion_state(
    project_root: Path | str,
    source_name: str,
    **updates: object,
) -> dict[str, object]:
    """Atomically persist one source's conversion progress."""
    path = _state_path(project_root, source_name)
    current = read_ingestion_state(project_root, source_name, has_documents=False)
    current.update(updates)
    current.update({
        "schema_version": 1,
        "source_path": f"data/{source_name}",
        "updated_at": _utc_now(),
    })
    current["progress"] = max(0, min(100, int(current.get("progress") or 0)))
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{os.getpid()}.{threading.get_ident()}.tmp")
    try:
        temporary.write_text(
            json.dumps(current, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
    return current


def mark_interrupted_ingestions(project_root: Path | str) -> int:
    """Turn stale `processing` markers into recoverable failed states on startup."""
    originals_root = Path(project_root).resolve() / ".knowledge_originals"
    if not originals_root.is_dir():
        return 0
    changed = 0
    for source_root in sorted(originals_root.iterdir()):
        if not source_root.is_dir() or source_root.name.startswith("."):
            continue
        state = read_ingestion_state(project_root, source_root.name, has_documents=False)
        if state.get("status") != "processing":
            continue
        owner_pid = int(state.get("owner_pid") or 0)
        if owner_pid and owner_pid != os.getpid() and _process_is_alive(owner_pid):
            continue
        cleanup_pending_outputs(project_root, source_root.name, state)
        write_ingestion_state(
            project_root,
            source_root.name,
            status="failed",
            stage="failed",
            error="服务在后处理完成前中断，请重新上传该批文件",
            finished_at=_utc_now(),
            pending_outputs=[],
            owner_pid=0,
        )
        changed += 1
    return changed


def cleanup_pending_outputs(
    project_root: Path | str,
    source_name: str,
    state: dict[str, object],
) -> int:
    """Remove only interrupted-batch files whose recorded digest still matches."""
    project_root = Path(project_root).resolve()
    allowed_roots = (
        (project_root / "data" / source_name).resolve(),
        (project_root / ".knowledge_originals" / source_name).resolve(),
    )
    removed = 0
    pending = state.get("pending_outputs")
    if not isinstance(pending, list):
        return removed
    for entry in pending:
        if not isinstance(entry, dict):
            continue
        relative = str(entry.get("path") or "").strip()
        expected_sha256 = str(entry.get("sha256") or "").strip().lower()
        if not relative or not re.fullmatch(r"[0-9a-f]{64}", expected_sha256):
            continue
        target = (project_root / relative).resolve()
        if not any(target == root or root in target.parents for root in allowed_roots):
            continue
        if not target.is_file():
            continue
        digest = hashlib.sha256()
        try:
            with target.open("rb") as stream:
                for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                    digest.update(chunk)
            if digest.hexdigest() != expected_sha256:
                continue
            target.unlink()
            removed += 1
        except OSError:
            continue
    return removed


def _lock_key(project_root: Path | str, source_name: str) -> str:
    return f"{Path(project_root).resolve()}::{source_name}"


def _operation_lock_path(project_root: Path | str, source_name: str) -> Path:
    project_root = Path(project_root).resolve()
    locks_root = (project_root / ".knowledge_locks").resolve()
    target = (locks_root / f"{source_name}.lock").resolve()
    if target.parent != locks_root or not source_name or source_name.startswith("."):
        raise ValueError("知识源锁路径无效")
    return target


def _process_is_alive(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    return True


def _acquire_process_lock(lock_path: Path, blocking: bool) -> str:
    token = f"{os.getpid()}:{threading.get_ident()}:{os.urandom(8).hex()}"
    while True:
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            descriptor = os.open(lock_path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        except FileExistsError:
            try:
                owner = lock_path.read_text(encoding="utf-8").split(":", 1)[0]
                owner_pid = int(owner)
            except (OSError, ValueError):
                owner_pid = 0
            try:
                lock_is_fresh = time_module.time() - lock_path.stat().st_mtime < 5
            except OSError:
                lock_is_fresh = False
            if (owner_pid and _process_is_alive(owner_pid)) or (not owner_pid and lock_is_fresh):
                if not blocking:
                    return ""
                time_module.sleep(0.05)
                continue
            try:
                lock_path.unlink()
            except FileNotFoundError:
                pass
            except OSError:
                if not blocking:
                    return ""
                time_module.sleep(0.05)
            continue
        try:
            os.write(descriptor, token.encode("ascii"))
        finally:
            os.close(descriptor)
        return token


def _release_process_lock(lock_path: Path, token: str) -> None:
    try:
        if lock_path.read_text(encoding="utf-8") == token:
            lock_path.unlink()
    except OSError:
        pass
    finally:
        with contextlib.suppress(OSError):
            lock_path.parent.rmdir()


@contextmanager
def source_operation_lock(
    project_root: Path | str,
    source_name: str,
    *,
    blocking: bool = False,
) -> Iterator[bool]:
    """Serialize upload, source mutation, and job snapshot operations per source."""
    key = _lock_key(project_root, source_name)
    with _SOURCE_LOCKS_GUARD:
        lock = _SOURCE_LOCKS.setdefault(key, threading.Lock())
    acquired = lock.acquire(blocking=blocking)
    lock_path = _operation_lock_path(project_root, source_name)
    try:
        process_token = _acquire_process_lock(lock_path, blocking) if acquired else ""
    except Exception:
        if acquired:
            lock.release()
        raise
    if acquired and not process_token:
        lock.release()
        acquired = False
    try:
        yield acquired
    finally:
        if acquired:
            _release_process_lock(lock_path, process_token)
            lock.release()


def _decode_text(raw: bytes) -> tuple[str, str]:
    if not raw:
        raise ValueError("不允许上传空文件")
    if b"\x00" in raw:
        raise ValueError("文件包含二进制内容，请检查文件格式")
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别文件编码，请转换为 UTF-8 文本后上传")


def _validate_office_archive(raw: bytes, filename: str) -> None:
    """Reject malformed/encrypted/oversized Office ZIP containers early."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_FILES:
                raise ValueError(f"{filename} 内部文件数量过多")
            if sum(entry.file_size for entry in entries) > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                raise ValueError(f"{filename} 解压后超过 100 MB")
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise ValueError(f"{filename} 已加密，暂不支持解析")
    except zipfile.BadZipFile as exc:
        raise ValueError(f"{filename} 不是有效的 Office 文件") from exc


def _normalize_newlines(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _finish_markdown(text: str, filename: str) -> str:
    text = _normalize_newlines(text).strip("\ufeff\n")
    if not re.sub(r"\s+", "", text):
        raise ValueError(f"{filename} 转换后没有可用文本")
    if len(text) > MAX_NORMALIZED_DOCUMENT_CHARS:
        raise ValueError(f"{filename} 转换后的文本超过 500 万字符，请拆分后上传")
    return text + "\n"


def _source_heading(filename: str) -> str:
    title = Path(filename).stem.replace("\n", " ").strip() or "知识文档"
    return f"# {title}"


def _markdown_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        text = value.isoformat()
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    else:
        text = str(value)
    return _normalize_newlines(text).replace("|", "\\|").replace("\n", "<br>").strip()


def _markdown_table(rows: list[list[object]]) -> str:
    cleaned: list[list[str]] = []
    width = 0
    for row in rows:
        cells = [_markdown_cell(value) for value in row]
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue
        cleaned.append(cells)
        width = max(width, len(cells))
    if not cleaned or width == 0:
        return ""

    padded = [row + [""] * (width - len(row)) for row in cleaned]
    header = padded[0]
    # Markdown requires a header row.  Empty source headers get stable labels,
    # while non-empty values are kept exactly as they appeared in the source.
    header = [value or f"列 {index + 1}" for index, value in enumerate(header)]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in padded[1:])
    return "\n".join(lines)


def _word_to_markdown(raw: bytes, filename: str) -> str:
    _validate_office_archive(raw, filename)
    try:
        from docx import Document
        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError as exc:  # pragma: no cover - exercised by deployment only
        raise ValueError("服务器缺少 Word 解析组件，请重新安装项目依赖") from exc

    try:
        document = Document(io.BytesIO(raw))
    except Exception as exc:
        raise ValueError(f"{filename} 不是有效的 DOCX 文件") from exc

    parts = [_source_heading(filename)]
    content_count = 0
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            paragraph = Paragraph(child, document)
            text = paragraph.text.strip()
            if not text:
                continue
            content_count += 1
            style = paragraph.style
            style_name = (getattr(style, "name", "") or "").strip()
            style_id = (getattr(style, "style_id", "") or "").strip()
            heading_match = re.search(r"(?:heading|标题)\s*([1-9])", f"{style_name} {style_id}", re.I)
            if style_name.casefold() == "title" or style_id.casefold() == "title":
                parts.append(f"## {text}")
            elif heading_match:
                level = min(6, int(heading_match.group(1)) + 1)
                parts.append(f"{'#' * level} {text}")
            elif paragraph._p.pPr is not None and paragraph._p.pPr.numPr is not None:
                parts.append(f"- {text}")
            else:
                parts.append(text)
        elif isinstance(child, CT_Tbl):
            table = Table(child, document)
            rendered = _markdown_table([
                [cell.text for cell in row.cells]
                for row in table.rows
            ])
            if rendered:
                parts.append(rendered)
                content_count += 1
    if not content_count:
        raise ValueError(f"{filename} 转换后没有可用文本")
    return _finish_markdown("\n\n".join(parts), filename)


def _excel_to_markdown(raw: bytes, filename: str) -> str:
    _validate_office_archive(raw, filename)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised by deployment only
        raise ValueError("服务器缺少 Excel 解析组件，请重新安装项目依赖") from exc

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    except Exception as exc:
        raise ValueError(f"{filename} 不是有效的 XLSX 文件") from exc

    parts = [_source_heading(filename)]
    cell_count = 0
    content_count = 0
    try:
        for sheet in workbook.worksheets:
            rows: list[list[object]] = []
            for row in sheet.iter_rows(values_only=True):
                cell_count += len(row)
                if cell_count > MAX_SPREADSHEET_CELLS:
                    raise ValueError(f"{filename} 超过 20 万个单元格，请拆分后上传")
                values = list(row)
                while values and values[-1] is None:
                    values.pop()
                if values and any(value is not None and str(value).strip() for value in values):
                    rows.append(values)
            rendered = _markdown_table(rows)
            if rendered:
                parts.append(f"## 工作表：{sheet.title}\n\n{rendered}")
                content_count += 1
    finally:
        workbook.close()
    if not content_count:
        raise ValueError(f"{filename} 转换后没有可用文本")
    return _finish_markdown("\n\n".join(parts), filename)


def _pdf_to_markdown(raw: bytes, filename: str) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - exercised by deployment only
        raise ValueError("服务器缺少 PDF 解析组件，请重新安装项目依赖") from exc

    try:
        reader = PdfReader(io.BytesIO(raw), strict=False)
        if reader.is_encrypted and not reader.decrypt(""):
            raise ValueError(f"{filename} 已加密，暂不支持解析")
        if len(reader.pages) > MAX_PDF_PAGES:
            raise ValueError(f"{filename} 超过 2000 页，请拆分后上传")
        page_texts = [
            _normalize_newlines(page.extract_text() or "").strip()
            for page in reader.pages
        ]
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"{filename} 不是有效的 PDF 文件") from exc

    extracted = re.sub(r"\s+", "", "".join(page_texts))
    if not extracted:
        raise ValueError(f"{filename} 未检测到文本层，可能是扫描版 PDF，请先进行 OCR 后再上传")

    parts = [_source_heading(filename)]
    for index, text in enumerate(page_texts, start=1):
        if text:
            parts.append(f"## 第 {index} 页\n\n{text}")
    return _finish_markdown("\n\n".join(parts), filename)


def normalize_knowledge_document(filename: str, raw: bytes) -> NormalizedKnowledge:
    """Convert one supported upload to the canonical Markdown representation."""
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_KNOWLEDGE_SUFFIXES:
        allowed = "、".join(sorted(SUPPORTED_KNOWLEDGE_SUFFIXES))
        raise ValueError(f"不支持 {suffix or '无扩展名'} 文件；支持：{allowed}")
    if not raw:
        raise ValueError("不允许上传空文件")

    if suffix in {".md", ".markdown", ".txt"}:
        text, encoding = _decode_text(raw)
        if suffix == ".txt":
            if not re.sub(r"\s+", "", text):
                raise ValueError(f"{filename} 转换后没有可用文本")
            text = f"{_source_heading(filename)}\n\n{text}"
        return NormalizedKnowledge(
            markdown=_finish_markdown(text, filename),
            source_format=suffix.removeprefix("."),
            source_encoding=encoding,
        )
    if suffix == ".docx":
        markdown = _word_to_markdown(raw, filename)
    elif suffix == ".xlsx":
        markdown = _excel_to_markdown(raw, filename)
    else:
        markdown = _pdf_to_markdown(raw, filename)
    return NormalizedKnowledge(
        markdown=markdown,
        source_format=suffix.removeprefix("."),
        source_encoding="binary",
    )
